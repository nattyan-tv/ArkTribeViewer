import os
import shutil
import sys
import json
import urllib.request
import urllib.error
import zipfile
import traceback
import subprocess
import openpyxl
import PySimpleGUI as sg

sg.theme('Default1')

PATH = sys.path[0].replace("\\", "/")
ARKTOOL = "https://github.com/Qowyn/ark-tools/releases/download/v0.6.4/ark-tools.zip"


def duplicate_rename2(file_path):
    # https://qiita.com/mareku/items/d29fc9bd46f40264d815
    if os.path.exists(file_path):
        name, ext = os.path.splitext(file_path)
        i = 1
        while True:
            # 数値を3桁などにしたい場合は({:0=3})とする
            new_name = "{} ({}){}".format(name, i, ext)
            if not os.path.exists(new_name):
                return new_name
            i += 1
    else:
        return file_path


def extractJson(source_path: str, destination_path: str) -> None:
    subprocess.run(
        [
            os.path.join(PATH, "ArkTools/ark-tools.exe"),
            "t2j",
            source_path,
            destination_path
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=os.path.join(PATH, "ArkTools")
    )


def scrapeTribes(path: str) -> list:
    tribe_datas = [
        i for i in os.listdir(path=path) if i.split(".")[-1] == "arktribe"
    ]
    if os.path.exists(os.path.join(PATH, "temp")):
        shutil.rmtree(os.path.join(PATH, "temp"))
    os.mkdir(os.path.join(PATH, "temp"))
    for i in tribe_datas:
        extractJson(
            os.path.join(path, i),
            os.path.join(PATH, f"temp/{'.'.join(i.split('.')[:-1])}.json")
        )
    return [
        i for i in os.listdir(path=os.path.join(PATH, "temp")) if i.split(".")[-1] == "json"
    ]


def scrapeMembers(path: str) -> dict:
    tribeData = json.load(
        open(os.path.join(PATH, f"temp/{path}"), "r", encoding="utf-8"))
    tribeName = tribeData["tribe"]["properties"][0]["value"][0]["value"]
    tribeMembersName = tribeData["tribe"]["properties"][0]["value"][3]["value"]
    tribeMembersID = tribeData["tribe"]["properties"][0]["value"][4]["value"]
    cellData = [["Name", "ID"]]
    for i in range(len(tribeMembersName)):
        cellData.append([tribeMembersName[i], tribeMembersID[i]])
    return {tribeName: cellData}


def getArkTools() -> None:
    if os.path.exists(os.path.join(PATH, "ArkTools")):
        shutil.rmtree(os.path.join(PATH, "ArkTools"))
    os.mkdir(os.path.join(PATH, "ArkTools"))
    try:
        with urllib.request.urlopen(ARKTOOL) as f:
            data = f.read()
            with open(os.path.join(PATH, "ArkTools", "download.zip"), mode="wb") as sf:
                sf.write(data)
        with zipfile.ZipFile(os.path.join(PATH, "ArkTools", "download.zip")) as z:
            z.extractall(os.path.join(PATH, "ArkTools"))
        sg.popup_ok(f"ツールの準備が完了しました。", title="完了", keep_on_top=True)
    except Exception:
        sg.popup_ok(
            f"ツールの展開中にエラーが発生しました。\n\n{traceback.format_exc()}",
            title="エラー",
            keep_on_top=True
        )


def createWorkbook(tribe_data: dict) -> str:
    if not os.path.exists(os.path.join(PATH, "export")):
        os.mkdir(os.path.join(PATH, "export"))
    BOOK = openpyxl.Workbook()
    INFO = BOOK["Sheet"]
    INFO.title = "INFO"
    for i in range(len(list(tribe_data.keys()))):
        INFO.cell(row=i+1, column=1).value = list(tribe_data.keys())[i]
    for name, members in tribe_data.items():
        BOOK.create_sheet(title=name)
        Sheet = BOOK[name]
        for j in range(len(members)):
            Sheet.cell(row=j+1, column=1).value = members[j][0]
            Sheet.cell(row=j+1, column=2).value = members[j][1]
    filename = duplicate_rename2(
        os.path.join(
            PATH, "export/export.xlsx"
        )
    )
    BOOK.save(filename=filename)
    # BOOK.close()
    return filename


layout = [
    [
        sg.MenuBar([['ファイル', ['ツール', ['インストール', 'アンインストール'], '閉じる']]])
    ],
    [
        sg.Text('セーブデータがあるフォルダ（ShooterGame/Saved/SavedArksLocal）を選択してください。')
    ],
    [
        sg.Text('フォルダ', size=(10, 1)), sg.Input(),
        sg.FolderBrowse('フォルダを選択', key='folder_path')
    ],
    [
        sg.Button('書き出し', key='save')
    ],
    [
        sg.Output(size=(80, 20))
    ]
]

window = sg.Window('ArkTribeViewer', layout)

while True:
    event, values = window.read()
    if event == 'インストール':
        getArkTools()

    if event == 'アンインストール':
        if os.path.exists(os.path.join(PATH, "ArkTools")):
            shutil.rmtree(os.path.join(PATH, "ArkTools"))
        sg.popup_ok(f"ツールを削除しました。", title="完了", keep_on_top=True)

    if event == sg.WIN_CLOSED or event == "閉じる":
        break

    if event == 'save':
        if not os.path.isfile(os.path.join(PATH, "ArkTools/ark-tools.exe")):
            print(f"書き込み失敗: ツールが見つかりませんでした。")
            sg.popup_ok(
                f"ツールが存在しません。\n画面上部のメニューからインストールしてください。",
                title="エラー",
                keep_on_top=True
            )
        else:
            tribes = {}
            for i in scrapeTribes(values["folder_path"]):
                dt = scrapeMembers(i)
                key = list(dt.keys())[0]
                value = list(dt.values())[0]
                if key in list(tribes.keys()):
                    while True:
                        inc = 1
                        key = key + f"({inc})"
                        if key in list(tribes.keys()):
                            inc += 1
                        else:
                            dt = {key: value}
                            break
                tribes.update(dt)

            path = createWorkbook(tribes)
            print(f"書き込み終了: {path}")
            sg.popup_ok(
                f"ファイルの書き込みに完了しました。\n\n・出力先\n{path}", title="完了", keep_on_top=True)

window.close()
